import openpyxl
from pattern.en import comparative, superlative,pluralize, singularize, conjugate
wb = openpyxl.load_workbook('NLP File_all.xlsx')
sheet = wb.get_sheet_by_name('NLP')
n_rows=sheet.max_row
n_cols=sheet.max_column

def combine(inputlist, result):
    if not inputlist:
        return result
    if not result:
        newresult=inputlist[0]
    else:
        newresult=[]
        for i in result:
            for j in inputlist[0]:
                newresult.append(i+' '+j)
    return combine(inputlist[1:], newresult)



data_dict=[]
for row in range(2,n_rows):
    phrase=[]
    for col in range(1,n_cols,2):
        word=sheet.cell(row=row,column=col).value
        try:
            if word:
                word=str(word).lower()
                order=sheet.cell(row=row,column=col+1).value
                order=int(order) if order else 1

                #for the noun
                if col==9 or col==11:
                    templist = []
                    singular = singularize(word)
                    templist.append(singular)
                    templist.append(pluralize(singular))
                    templist = list(set(templist))

                #for the verb
                elif col==13:
                    templist=[]
                    templist.append(conjugate(word,'inf'))
                    templist.append(conjugate(word,'1sg'))
                    templist.append(conjugate(word,'2sg'))
                    templist.append(conjugate(word,'3sg'))
                    templist.append(conjugate(word,'pl'))
                    templist.append(conjugate(word,'part'))
                    templist.append(conjugate(word,'p'))
                    templist.append(conjugate(word,'1sgp'))
                    templist.append(conjugate(word,'2sgp'))
                    templist.append(conjugate(word,'3gp'))
                    templist.append(conjugate(word,'ppl'))
                    templist.append(conjugate(word,'ppart'))
                    templist = list(set(templist))
                    templist = filter(None,templist)
                    templist=[str(x) for x in templist]

                else: templist=[word]
                phrase.insert(order-1,templist)

        except ValueError:
            pass

    data_dict+=combine(phrase,[])
    data_dict=list(set(data_dict))
    data_dict=filter(None,data_dict)
    data_dict.sort()
#print data_dict
txtfile = open('variation_all.txt', 'w')
for item in data_dict:
  txtfile.write("%s\n" % item)
txtfile.close()
